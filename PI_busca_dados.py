from os import error
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import datetime
import openpyxl            # Módulo para trabalhar com tempo e conversões
import pymysql.cursors        # Módulo Mysql (Banco de Dados)
import pymysql
import numpy as np
import csv
import random
from openpyxl import Workbook
import string
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.comments import Comment

################################ FUNÇÕES MYSQL ###############################
# Insere mensagens recebidas no banco de dados.
# Inicia conexão com o Banco de Dados Mysql
try:
    connection = pymysql.connect(host='localhost',
                             user='root',
                             password='',
                             db='testepi',
                             charset='utf8mb4',
                             cursorclass=pymysql.cursors.DictCursor)
except:
    print("Não foi possível conectar-se ao banco de dados.")

wb = Workbook()
wb = openpyxl.load_workbook('C:\\Users\\Guilh\\OneDrive\\Documentos\\NetBeansProjects\\GUI_PI\\bases\\Economia.xlsx')
std = wb['Dados']
wb.remove(std)
sheet_plot =wb.create_sheet(0)
sheet_plot.title="Dados"
c = connection.cursor()

sql = "SELECT * FROM datas;"

corrente = []
x = []
data = []
row = []
current = []
hour = []
separador = []
lastread = 0.0
funcionamento = 0
timeh = 0
timem = 0

def read_sql():
    i = 0
    c.execute(sql)
    for row in c.fetchall():
        for current, diff_hour in row.items():
            temp = [current, diff_hour]
            x.append(temp)
    
    for line in x:
        for leitura in line:
            i += 1            
            position = 'A' + str(i)
            sheet_plot[position] = str(leitura)

def receive_xl():
    ccurrent = 4
    chour = 10
    while sheet_plot['A' + str(ccurrent)].value != None:
        current.append(float(sheet_plot['A' + str(ccurrent)].value))
        hour.append(sheet_plot['A' + str(chour)].value)       
        ccurrent += 10
        chour += 10
        lastread = current
        funcionamento = hour
    lastread =sheet_plot['A' + str(ccurrent-10)].value
    timeh = int((chour-10)/36000)
    chour -= 36010
    timem = int(chour / 600)
    funcionamento = str(timeh) + "h" + str(timem) + "min"
    return [lastread,funcionamento]


def gera_txt():
    arquivo = open("C:\\Users\\Guilh\\OneDrive\\Documentos\\NetBeansProjects\\GUI_PI\\bases\\estatisticas.txt","w")
    arquivo.write(str(separador[0]) + "\n")
    arquivo.write(str(separador[1]) + "\n")
    
read_sql()
separador = receive_xl()
# Salva a planilha em formato xlsx
wb.save('Economia.xlsx')
gera_txt()


